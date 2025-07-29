
# 1. Vytvorenie letov pomocou namedtuple nazov Let a atributy cislo, odlet, ciel, pocet_pasazierov

from collections import namedtuple

Let = namedtuple("Let", ["cislo", "odlet", "ciel", "pocet_pasazierov"])

lety = [Let("FR007", "Viedeň", "Chania", 154),
        Let("FR017", "BUD", "MIL", 124),
        Let("FR056", "Bratislava", "Zakythos", 165)]

# 2. Prevod na dataclass, pridanie typu lietadla

from dataclasses import dataclass
import random

@dataclass
class LetClass:
    cislo: str
    odlet: str
    ciel: str
    pocet_pasazierov: int
    typ_lietadla: str

data_lety = [LetClass(*let, random.choice(["Boeing", "Airbus", "JetSky"])) for let in lety]

# 3. Uloženie do Excelu (openpyxl)

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.append(["Cislo", "Odlet", "Ciel", "Pocet pasazierov", "Typ lietadla"])

for let in data_lety:
    ws.append([let.cislo, let.odlet, let.ciel, let.pocet_pasazierov, let.typ_lietadla])

wb.save("letecka_preprava_vystup.xlsx")


# 4. Načítanie z Excelu
from openpyxl import load_workbook
wb = load_workbook("letecka_preprava_vystup.xlsx")
ws = wb.active 

# Prečítaj všetky riadky z Excelu a zapis do tablib Dataset a exportuj do CSV

import tablib

data = tablib.Dataset()
rows = list(ws.iter_rows(values_only=True))
data.headers = rows[0]

for row in rows[1:]:
    data.append(row)

with open("lety.csv", "w", encoding="UTF-8", newline="") as f:
    f.write(data.export("csv"))


# 5. Sčítanie pasažierov z CSV cez tablib

number = sum([r["Pocet pasazierov"] for r in data.dict])
print(f"Pocet pasazierov je {number}")